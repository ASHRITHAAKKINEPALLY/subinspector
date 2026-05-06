import os
import re
import io
import base64
import asyncio
import traceback
import httpx
import pdfplumber
import openpyxl
from docx import Document

GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
CLICKUP_API_KEY = os.environ.get("CLICKUP_API_KEY")

# ── Scope configuration ───────────────────────────────────────────────────────
# Three levels of matching (checked in order, OR logic):
#   1. folder.id  — matches tasks in a specific ClickUp folder
#   2. list.id    — fallback for tasks whose folder.id = "none" (list directly in space)
#   3. space.id   — broadest: matches ANY task (including master tickets) in the space
#
# ENFORCEMENT = full gate enforcement + revert (IH only)
# ADVISORY    = comment-only gate checks, no revert (external clients)
#
# SPACES are the most important for master tickets: master/epic tickets often
# live in a "Backlogs" or "Master Tickets" list that has a different folder.id
# than sprint lists. Setting the space ID catches them automatically.

ENFORCEMENT_FOLDERS = os.environ.get("ENFORCEMENT_FOLDERS", "90165998786").split(",")
ENFORCEMENT_SPACES  = [x.strip() for x in os.environ.get("ENFORCEMENT_SPACES", "").split(",") if x.strip()]

# Client folders/spaces for advisory mode (comment only, no status changes).
_DEFAULT_ADVISORY_FOLDERS = ",".join([
    "90161200308",  # HexClad
    "90161875051",  # Saxx
    "90169023555",  # Bboutique
    "90167972037",  # Naked & Thriving (N&T)
    "90169078001",  # Javvy Coffee
    "90164305799",  # Yum Brands
    "90160230070",  # Momentous Projects
    "90020845754",  # BPN - BarePerformanceNutrition (Consulting)
    "90160770330",  # BPN (DE)
])
ADVISORY_FOLDERS = os.environ.get("ADVISORY_FOLDERS", _DEFAULT_ADVISORY_FOLDERS).split(",")
ADVISORY_SPACES  = [x.strip() for x in os.environ.get("ADVISORY_SPACES", "").split(",") if x.strip()]

print(f"[AGENT] Startup check — GROQ_API_KEY={'SET (' + GROQ_API_KEY[:8] + '...)' if GROQ_API_KEY else 'MISSING ⚠️'}", flush=True)
print(f"[AGENT] Startup check — CLICKUP_API_KEY={'SET (' + CLICKUP_API_KEY[:8] + '...)' if CLICKUP_API_KEY else 'MISSING ⚠️'}", flush=True)
print(f"[AGENT] Startup check — ENFORCEMENT_FOLDERS={ENFORCEMENT_FOLDERS}", flush=True)
print(f"[AGENT] Startup check — ENFORCEMENT_SPACES={ENFORCEMENT_SPACES or '(not set — add via HF secret to catch master tickets)'}", flush=True)
print(f"[AGENT] Startup check — ADVISORY_FOLDERS={ADVISORY_FOLDERS}", flush=True)
print(f"[AGENT] Startup check — ADVISORY_SPACES={ADVISORY_SPACES or '(not set — add via HF secret to catch master tickets)'}", flush=True)

GROQ_URL = "https://api.groq.com/openai/v1/chat/completions"
CLICKUP_BASE = "https://api.clickup.com/api/v2"

PRE_EXEC_STATUSES = ["ready", "in progress", "in progess", "development", "code-review", "code review"]
CLOSURE_STATUSES = ["qa", "uat", "prod review", "prod-review", "complete", "done", "ready to close"]

# When the webhook payload is missing previous_status, use these maps to
# determine where to revert. CLOSURE reverts to prod-review; PRE-EXECUTION
# reverts to "open" (the natural state before work begins).
CLOSURE_REVERT_MAP = {
    "complete":       "prod-review",
    "done":           "prod-review",
    "ready to close": "prod-review",
    # Tickets already at prod-review/uat/qa are already under review —
    # don't cascade them lower, just post the FAIL comment.
}
PRE_EXEC_REVERT_MAP = {
    "ready":        "open",
    "in progress":  "open",
    "in progess":   "open",
    "development":  "open",
    "code-review":  "open",
    "code review":  "open",
}

# Trigger patterns — require a leading slash so the bot's own next-steps
# instructions never match and cause a loop.
# [ \t\xa0]+ covers regular space, tab, and non-breaking space (U+00A0)
# which ClickUp inserts when typing in the comment box.
_TRIGGER_PATTERNS = [
    re.compile(r'/subinspector[ \t\xa0]+check\b', re.IGNORECASE),
    re.compile(r'/si[ \t\xa0]+check\b',           re.IGNORECASE),
]

# User ID of the account the bot posts under — skip comments from this user
# to prevent the bot from reacting to its own comments.
BOT_USER_ID = os.environ.get("BOT_USER_ID", "100965864")

def _is_trigger(text: str) -> bool:
    return any(p.search(text) for p in _TRIGGER_PATTERNS)

# Shared context sent for every gate evaluation (~600 tokens)
_SYSTEM_COMMON = """You are SubInspector, a ClickUp ticket quality gate enforcer for the Instant Hydration (IH) DE team.
Evaluate tickets against a 6-point gate checklist. Score each check PASS or FAIL. Pass = 6/6. Never use subjective phrasing.

CRITICAL — IGNORE BOT COMMENTS: Any comment that starts with "🤖 **SubInspector" is an automated bot report from a previous run. DO NOT use these as evidence for any check. DO NOT let prior PASS/FAIL results influence your evaluation. Judge only the ticket description, human comments, and attachments.

TIER CLASSIFICATION (determine first):
- T1: Label fix, filter change, config tweak. Light gate — description + success criteria sufficient.
- T2: Analysis, moderate modeling, enhancement. All 6 BA Inputs required.
- T3: Title/description contains "dashboard/dataset/model/client/logic/allocation" + new build. Full gate, any TBD = FAIL.
If comment/description contains "Tier: T1/T2/T3", use that tier.

SIMPLE DISCREPANCY BUGS (Bug Category = Logic Misalignment OR title has "discrepancy/mismatch/incorrect"):
Minimum INTAKE requires only: (1) plain description of mismatch, (2) affected metric named, (3) one validation check.

KEY PEOPLE — do NOT count as DE assignees: Komal Saraogi (PM/BA), Frido (management). Anudeep = valid only for BI tickets.
BIGQUERY: Full path required: project.dataset.table. "in BQ" or no path = FAIL.
BI TICKET: Treat as BI if title/description references Tableau, Power BI, dashboards, PBIX, workbooks.
Missing sections = FAIL. Placeholder text ("TBD", "N/A to fill later", "will update") = FAIL.

MASTER TICKET: If ticket has subtasks OR title contains master/epic/initiative/tracker/rollout, append after gate checks:
MASTER TICKET: YES
SCOPE ITEMS FOUND: [list from description]
SUBTASK COVERAGE: | Scope Item | Covered By | Status |
SCOPE VERDICT: FULLY COVERED / PARTIALLY COVERED / GAPS FOUND

RESPONSE FORMAT (strict):
TIER: [T1/T2/T3] — [reason]
RESULT: PASS or FAIL
SCORE: X/6
CHECKS:
| # | Check | Result | Detail |
|---|-------|--------|--------|
| 1 | [name] | ✅ PASS or ❌ FAIL | [one-line — for FAIL: what is missing and what to add] |
| 2 | [name] | ✅ PASS or ❌ FAIL | [detail] |
| 3 | [name] | ✅ PASS or ❌ FAIL | [detail] |
| 4 | [name] | ✅ PASS or ❌ FAIL | [detail] |
| 5 | [name] | ✅ PASS or ❌ FAIL | [detail] |
| 6 | [name] | ✅ PASS or ❌ FAIL | [detail] |
SUMMARY: [one sentence verdict + most critical gap if FAIL]"""

# Gate-specific checklists (~250 tokens each) — only the relevant one is sent
_GATE_CHECKS = {
    "INTAKE": """
INTAKE GATE — Generic (6 checks):
1. Title-Description Coherence — title and problem statement describe the SAME work. PASS if they are about the same general area, entity, or goal even if worded differently or at different levels of detail. FAIL ONLY if the title and description are clearly about two different things (e.g. title says "revenue dashboard" but description is about user churn pipeline).
2. Steps to Reproduce / Context — new person can understand without a meeting: navigation, filters, date range, context.
3. Definition of Done — explicit, observable end state stated. FAIL if vague ("fix it") or unmeasurable.
4. Screenshots/Evidence — INTAKE ONLY RULE: work has NOT started yet, so finished output does NOT exist. PASS if any of these are present: (a) screenshot of the current/broken state showing the problem, (b) description or sample of expected output format (table columns, sample rows, metric name + formula), (c) mockup or wireframe. FAIL ONLY if the ticket makes a claim about wrong/missing data or a UI issue AND has absolutely no screenshot, no format description, and no example of any kind. PLANNING/INITIATIVE EXCEPTION: If the ticket is a planning, strategy, or initiative ticket where the deliverable IS a new artifact to be created (deck, document, presentation, plan, design spec, narrative), PASS this check automatically — no evidence of a non-existent artifact can exist at intake.
5. Mandatory Fields — Problem Statement, Expected Output, Definition of Done, Data Source all present and non-empty with substantive content. PASS if the BA sections contain clear intent and value even if not formatted as a strict user-story template — do not fail for phrasing style when substance is present.
6. DE Actionability — expected output clear, BQ path present if DE work is in scope, no TBDs. Actionable without a meeting. PASS AUTOMATICALLY when the ticket involves no data engineering work (no BQ query, no pipeline, no table build, no ingestion, no transformation, no SQL). When DE work IS in scope: scan the entire description for any string matching project.dataset.table (e.g. pulse-instanthydration.dataset.tablename or project.dataset.table_*). PASS if at least one such path is found. Do NOT fail because paths are labeled "proposed"/"target" or have wildcard suffixes — new-build tickets provide target paths before the table exists, and that is acceptable. FAIL only when DE work is explicitly in scope AND no BQ path of any kind appears anywhere in the description.

INTAKE GATE — BI Tickets (use instead of Generic when BI ticket detected):
1. Problem Statement names dashboard, target persona, and business value.
2. BI Tool explicitly specified (Power BI / Tableau + workspace/embed target).
3. BigQuery path present — scan the entire description for any string matching the pattern project.dataset.table (dots between three non-space segments, e.g. pulse-instanthydration.instanthydration_4927_prod_raw.Northbeam_Ads_data_*). PASS if at least one such path is found anywhere. FAIL only if no BQ path of any kind exists in the description. Do NOT fail because paths are labeled "proposed" or "target" or have wildcard suffixes — those are acceptable.
4. KPIs/Metrics defined with calculation logic or spec/BRD reference.
5. Definition of Done — what the finished dashboard shows and how sign-off is given.
6. Screenshot/Mockup/Wireframe — PASS if mockup, wireframe, sample layout description, or screenshot of an existing similar report is attached or described. FAIL only if there is absolutely no visual reference or output format description of any kind.""",

    "PRE-EXECUTION": """
PRE-EXECUTION GATE — Generic (6 checks):
1. BA Inputs Complete — all 6 present: problem statement, expected output, scope/edge cases+timeline, validation checks, success criteria, data source+business context. FAIL if any missing or TBD.
2. Valid DE Assignee — at least one DE person assigned (not Komal/Frido; Anudeep only for BI). FAIL if only PM/BA assigned.
3. BigQuery path present — scan the description for any project.dataset.table string. PASS if found (proposed/target/wildcard paths are fine). FAIL only if no BQ path exists anywhere in the description.
4. Feasibility Assessment — technical review comment exists for T2/T3. PASS for T1 if self-evident. FAIL if absent for complex work.
5. Dependencies Identified and Unblocked — all dependencies recorded with owners and unblocked.
6. Scope Locked — no TBD/placeholder language in any execution-critical aspect.

PRE-EXECUTION GATE — BI Tickets (use instead of Generic when BI ticket detected):
1. All 6 BI Intake inputs complete — none TBD.
2. Valid BI developer assigned (Anudeep counts; PM/BA do not).
3. Granularity and filters defined (date range, drill-downs, slicers, row-level security).
4. Refresh cadence confirmed (live/daily/weekly/manual).
5. Upstream DE dependencies unblocked (source tables ready in BQ).
6. Scope locked — zero TBD in any metric, layout, or filter definition.""",

    "CLOSURE": """
CLOSURE GATE — Generic (6 checks):
IMPORTANT EVIDENCE RULE: Any Google Sheets URL, Google Docs URL, ClickUp Doc link, or GitHub PR link found in comments counts as attached evidence — PASS evidence checks even if the content is not directly readable. "Links found in comments" section lists these explicitly.
IMPORTANT SIGN-OFF RULE: A closure notes comment from Ashritha Akkinepally (team lead) IS the QA sign-off — always PASS check 3 when she has posted closure notes. A comment from any person other than the primary work assignee that confirms completion also counts.
IMPORTANT DOCS RULE: If the ticket is a bug fix, logic update, or config change with no mention of a BRD or downstream doc needing update, PASS check 6 — documentation N/A is implied by scope. Additionally, if the ticket title contains ANY of these words: mismatch, discrepancy, gap, fix, bug, logic, validation, incorrect, wrong — treat it as a logic/bug fix and PASS check 6 automatically regardless of what the comments say.
IMPORTANT COMPLETION RULE: Any comment containing phrases like "Moving ticket to Done", "Moving to Done", "Done", "Complete", "Completed", "work is done", "ticket is done", or any done/complete statement accompanied by 🎉 or ✅ counts as a valid completion confirmation. Such a comment PASSES check 1 AND check 5 automatically — do NOT require order-specific validation or explicit @mentions when this kind of statement exists.
IMPORTANT STAKEHOLDER RULE: If the ticket assignee, team lead (Ashritha Akkinepally), or any team member posts any comment confirming completion — PASS check 5. An explicit @mention is NOT required. Only FAIL check 5 if literally no team member has acknowledged the work is done in any comment.

1. Acceptance Criteria Addressed — PASS if any closing note, "Moving to Done" comment, completion comment, or SI auto-generated note confirms the work is done. FAIL only if there are literally zero completion confirmations of any kind in the comments. Do NOT require order-specific or metric-specific validation — a general "work done" statement is sufficient.
2. Evidence Attached — screenshots, query results, validation sheet links (Google Sheets/Docs), or before/after SQL/outputs attached, linked, or referenced. FAIL only if NO evidence of any kind exists anywhere.
3. QA Sign-Off — Ashritha Akkinepally's closure notes = sign-off (always PASS). Any comment from someone other than the primary work assignee confirming the work counts. FAIL only if nobody other than the sole assignee has weighed in at all.
4. No Open Subtasks — all subtasks closed (done/complete) or explicitly marked N/A.
5. Stakeholder Notified — PASS if any team member (assignee, team lead, or anyone) has posted a comment confirming the work is done, even without an explicit @mention. A "Moving ticket to Done 🎉" or equivalent statement counts. FAIL only if there is literally no completion acknowledgment from any team member.
6. Documentation Updated — updated, linked, or explicitly marked N/A. For bug fixes / logic updates / config changes, or any ticket whose title contains mismatch/discrepancy/gap/fix/bug/logic/validation/incorrect/wrong — PASS automatically (documentation N/A implied by scope). Only FAIL if the ticket is clearly a new feature or dashboard build with no documentation at all.

CLOSURE GATE — BI Tickets (use instead of Generic when BI ticket detected):
1. All KPIs validated with before/after numbers or screenshots.
2. Published dashboard link or final screenshot attached.
3. Stakeholder/client sign-off confirmed in a comment.
4. All subtasks closed or marked N/A.
5. Source tables/views documented in ticket or linked doc.
6. Publish and access handoff confirmed (right workspace, right users have access)."""
}


def get_system_prompt(gate: str) -> str:
    """Return a gate-specific system prompt (~850 tokens vs 2700 for the monolithic version)."""
    checks = _GATE_CHECKS.get(gate, _GATE_CHECKS["INTAKE"])
    return _SYSTEM_COMMON + "\n" + checks


# Regex that matches a BigQuery path: project.dataset.table
# project = lowercase + hyphens (e.g. pulse-instanthydration)
# dataset = lowercase + underscores + digits (e.g. instanthydration_4927_prod_raw)
# table   = any case + underscores + optional wildcard (e.g. Northbeam_Ads_data_*)
_BQ_PATH_RE = re.compile(
    r'\b[a-z][a-z0-9-]*\.[a-z][a-z0-9_]+\.[a-zA-Z][a-zA-Z0-9_*-]+\b'
)


def _fix_bq_check_false_fail(content: str, description: str) -> str:
    """Post-process LLM output: if check #3 is ❌ FAIL but a BQ path
    (project.dataset.table) is demonstrably present in the ticket description,
    override it to ✅ PASS.

    This handles the LLM's persistent tendency to fail check #3 on
    new-build/ingestion tickets even when full target BQ paths are present
    (e.g. Northbeam ingestion ticket with pulse-instanthydration.dataset.table).
    """
    # Only act when a FAIL exists and a BQ path is actually in the description
    if '❌ FAIL' not in content:
        return content
    bq_match = _BQ_PATH_RE.search(description or "")
    if not bq_match:
        return content  # no BQ path at all — FAIL is legitimate

    # Flip check #3 row from ❌ FAIL → ✅ PASS
    def _flip(m):
        name_col   = m.group(1)   # "| 3 | <check name> |"
        detail_col = m.group(2)   # "| <old detail> |"
        bq_found   = bq_match.group(0)[:70]
        return (
            f"{name_col} ✅ PASS "
            f"| BQ path confirmed in description (new-build target path acceptable): {bq_found} |"
        )

    new_content = re.sub(
        r'(\| 3 \|[^|]*\|)\s*❌ FAIL\s*(\|[^|]*\|)',
        _flip,
        content,
        count=1
    )
    if new_content != content:
        print(
            f"[AGENT] BQ path auto-override: check #3 → PASS "
            f"(found '{bq_match.group(0)[:60]}' in description)",
            flush=True
        )
    return new_content


# Threshold: table-embeds with more than this many rows are summarised
# (they're reference tables — column lists, lookup tables, etc. — not evidence).
# Small tables (≤ threshold) are formatted as readable text so the LLM can
# extract BQ paths, KPI definitions, etc. that live inside them.
_TABLE_EMBED_LARGE_ROWS = 10

def _process_table_embeds(text: str) -> str:
    """Transform [table-embed:...] blocks before they reach the LLM.

    • Large tables  (> _TABLE_EMBED_LARGE_ROWS rows): collapsed to a one-line
      token so they don't eat the description character budget.
    • Small tables  (≤ threshold): formatted as human-readable bullet rows so
      the LLM can read BQ paths, KPI specs, etc. embedded in them.
    """
    def _process(m):
        content = m.group(1)
        # Parse "R:C value" cells separated by " | "
        cells = {}
        for part in content.split(" | "):
            part = part.strip()
            cm = re.match(r'^(\d+):(\d+)\s+(.*)', part)
            if cm:
                r, c, val = int(cm.group(1)), int(cm.group(2)), cm.group(3).strip()
                cells[(r, c)] = val

        if not cells:
            return m.group(0)  # unparseable — leave as-is

        max_row = max(r for r, _ in cells)
        max_col = max(c for _, c in cells)

        # Large table → compact summary token
        if max_row > _TABLE_EMBED_LARGE_ROWS:
            return f"[table: {max_row} rows × {max_col} cols — reference table present ✓]"

        # Small table → format as readable bullet rows
        headers = [cells.get((1, c), f"col{c}") for c in range(1, max_col + 1)]
        lines = []
        for r in range(2, max_row + 1):
            row_vals = [cells.get((r, c), "") for c in range(1, max_col + 1)]
            pairs = [f"{h}: {v}" for h, v in zip(headers, row_vals) if v]
            if pairs:
                lines.append("• " + " | ".join(pairs))
        return "\n".join(lines) if lines else "[table: empty]"

    # Use a manual scanner instead of a greedy/lazy regex so that ']' characters
    # inside cell values (e.g. BQ array syntax or notes) don't truncate the match.
    # Strategy: after '[table-embed:' find the LAST ']' on the same logical block
    # by scanning forward until the next '[table-embed:' or a ']' that is followed
    # by whitespace / '[' / end-of-string (the natural block terminator pattern).
    result = []
    i = 0
    marker = "[table-embed:"
    while i < len(text):
        start = text.find(marker, i)
        if start == -1:
            result.append(text[i:])
            break
        result.append(text[i:start])
        # Scan for the closing ']' — prefer the one immediately before the next
        # '[table-embed:' or end-of-string so we capture the full block.
        content_start = start + len(marker)
        next_block = text.find(marker, content_start)
        search_end = next_block if next_block != -1 else len(text)
        # Walk backwards from search_end to find the last ']' in this block
        close = text.rfind("]", content_start, search_end)
        if close == -1:
            # No closing bracket found — emit verbatim and stop processing
            result.append(text[start:])
            i = len(text)
        else:
            content = text[content_start:close]
            # Build a fake match-like object so _process can use m.group(1) / m.group(0)
            class _M:
                def __init__(self, g0, g1):
                    self._g = [g0, g1]
                def group(self, n):
                    return self._g[n]
            result.append(_process(_M(text[start:close + 1], content)))
            i = close + 1
    return "".join(result)


def extract_comment_text(obj):
    """Pull plain text AND embedded URLs from any ClickUp comment object.

    ClickUp stores rich-text embeds (Google Sheets, Drive files, ClickUp Docs)
    as bookmark/embed blocks whose URL lives in attrs.href / attrs.url / attrs.link —
    NOT in the plain `text` field.  We must scan blocks even when text_content
    already has a value, because text_content never contains bookmark URLs.
    """
    if not obj:
        return ""

    # --- 1. Plain text fields -------------------------------------------------
    plain_text = ""
    for field in ("comment_text", "text_content", "text"):
        val = obj.get(field, "")
        if val and isinstance(val, str) and val.strip():
            plain_text = val.strip()
            break

    # --- 2. Rich-text block scanning (always run even if plain_text found) ----
    blocks = obj.get("comment") or []
    # ClickUp sometimes nests the comment object as a dict under the "comment" key
    # (e.g. webhook history item: {"comment": {"comment_text": "...", "id": "..."}}).
    # Recurse into it to extract text; don't treat it as a block list.
    if isinstance(blocks, dict):
        nested = extract_comment_text(blocks)
        if nested and not plain_text:
            plain_text = nested
        blocks = []  # no block-list to scan further
    block_parts = []
    block_urls  = []
    if isinstance(blocks, list):
        for b in blocks:
            if not isinstance(b, dict):
                continue
            # Plain text inside the block
            if b.get("text"):
                block_parts.append(b["text"].strip())
            # Bookmark / embed URL (Google Sheets, Drive, ClickUp Docs, etc.)
            # ClickUp uses both "attributes" and "attrs" depending on version
            for attr_key in ("attributes", "attrs"):
                attrs = b.get(attr_key) or {}
                if isinstance(attrs, dict):
                    for url_key in ("link", "href", "url"):
                        url_val = attrs.get(url_key, "")
                        if url_val and isinstance(url_val, str) and url_val.startswith("http"):
                            block_urls.append(url_val)
                            break  # only take first URL per attr dict
            # URL directly on the block object
            for url_key in ("url", "link", "href"):
                url_val = b.get(url_key, "")
                if url_val and isinstance(url_val, str) and url_val.startswith("http"):
                    block_urls.append(url_val)
                    break

    # --- 3. Combine -----------------------------------------------------------
    parts = []
    if plain_text:
        parts.append(plain_text)
    if block_parts:
        joined = " ".join(block_parts).strip()
        if joined and joined != plain_text:
            parts.append(joined)
    if block_urls:
        parts.append(" ".join(block_urls))

    result = " ".join(parts).strip()
    if result:
        return result

    return ""


def determine_gate(event, status, history_items):
    """Returns (gate, is_dry_run, trigger_comment_id, tier_override)."""
    status = status.lower()

    # Auto-check every new ticket at creation time (no revert — intake only)
    if event == "taskCreated":
        return "INTAKE", False, None, None

    # Status change → enforce gate, revert if it fails (is_dry_run=False)
    # IMPORTANT: check CLOSURE first — "ready to close" contains "ready" (a PRE-EXEC
    # substring) so checking PRE-EXEC first would mis-route it to the wrong gate.
    if event == "taskStatusUpdated":
        if any(s in status for s in CLOSURE_STATUSES):
            return "CLOSURE", False, None, None
        elif any(s in status for s in PRE_EXEC_STATUSES):
            return "PRE-EXECUTION", False, None, None

    # Manual si check comment → report only, never revert (is_dry_run=True)
    if event == "taskCommentPosted":
        trigger_comment_id = None
        comment_text = ""
        if history_items:
            item = history_items[0]
            # ClickUp may nest the comment object under "comment" or "data.comment"
            comment_obj = (
                item.get("comment")
                or (item.get("data") or {}).get("comment")
                or {}
            )
            own_id    = (comment_obj.get("id") if isinstance(comment_obj, dict) else None) or item.get("id") or None
            parent_id = (comment_obj.get("parent") if isinstance(comment_obj, dict) else None) or None
            # If si check was posted as a sub-comment (reply), reply to the
            # parent thread so the response appears in the same thread.
            # If it was a top-level comment, reply directly to that comment.
            trigger_comment_id = parent_id or own_id
            comment_text = extract_comment_text(comment_obj) or extract_comment_text(item) or ""
            print(f"[AGENT] determine_gate: own_id={own_id} parent_id={parent_id} → reply_to={trigger_comment_id} | text={repr(comment_text[:80])}", flush=True)

        tier_override = None
        tier_match = re.search(r"tier\s*:\s*(T[123])", comment_text, re.IGNORECASE)
        if tier_match:
            tier_override = tier_match.group(1).upper()

        if _is_trigger(comment_text):
            # IMPORTANT: check CLOSURE first — "ready to close" contains "ready"
            # (a PRE-EXEC substring), so checking PRE-EXEC first would mis-route it.
            if any(s in status for s in CLOSURE_STATUSES):
                # If the ticket is already at a terminal "done" state, just report —
                # never revert a ticket that's already been marked complete/done.
                # Enforcement (revert on fail) only applies to non-final closure
                # statuses like prod-review, uat, qa.
                already_done = status in ("complete", "done")
                return "CLOSURE", already_done, trigger_comment_id, tier_override
            elif any(s in status for s in PRE_EXEC_STATUSES):
                return "PRE-EXECUTION", True, trigger_comment_id, tier_override
            else:
                return "INTAKE", True, trigger_comment_id, tier_override
        else:
            print(f"[AGENT] determine_gate: no trigger found in comment text={repr(comment_text[:80])}", flush=True)

    return None, False, None, None


async def fetch_task(task_id):
    for attempt in range(3):
        try:
            async with httpx.AsyncClient(timeout=20) as client:
                response = await client.get(
                    f"{CLICKUP_BASE}/task/{task_id}",
                    headers={"Authorization": CLICKUP_API_KEY},
                    params={"include_subtasks": "true"}
                )
                raw = response.text
                if not raw or not raw.strip():
                    print(f"[AGENT] fetch_task attempt {attempt+1}: empty response (HTTP {response.status_code})", flush=True)
                    if attempt < 2:
                        await asyncio.sleep(3)
                        continue
                    raise ValueError(f"ClickUp returned empty response after 3 attempts (HTTP {response.status_code})")
                try:
                    data = response.json()
                except Exception:
                    raise ValueError(f"ClickUp non-JSON response (HTTP {response.status_code}): {raw[:200]}")
                if response.status_code >= 400 or "err" in data or not data.get("id"):
                    raise ValueError(f"ClickUp task fetch failed (HTTP {response.status_code}): {str(data)[:300]}")
                return data
        except (httpx.TimeoutException, httpx.ConnectError) as e:
            print(f"[AGENT] fetch_task network error attempt {attempt+1}: {e}", flush=True)
            if attempt < 2:
                await asyncio.sleep(3)
            else:
                raise
    raise ValueError(f"fetch_task failed after 3 attempts for {task_id}")


async def read_attachment(url, filename):
    """Download an attachment and extract its text content."""
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            # ClickUp attachment URLs are pre-signed CDN links — try without auth first
            resp = await client.get(url, follow_redirects=True)
            if resp.status_code in (401, 403):
                resp = await client.get(url, headers={"Authorization": CLICKUP_API_KEY}, follow_redirects=True)
            print(f"[AGENT] Download {filename}: status={resp.status_code} size={len(resp.content)}", flush=True)
            raw = resp.content
            ext = filename.lower().split(".")[-1] if "." in filename else ""

            # PDF
            if ext == "pdf":
                with pdfplumber.open(io.BytesIO(raw)) as pdf:
                    text = "\n".join(p.extract_text() or "" for p in pdf.pages)
                return f"[PDF: {filename}]\n{text[:3000]}"

            # Excel
            elif ext in ("xlsx", "xls"):
                wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
                lines = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    lines.append(f"Sheet: {sheet}")
                    for row in ws.iter_rows(values_only=True):
                        row_str = " | ".join(str(c) if c is not None else "" for c in row)
                        if row_str.strip(" |"):
                            lines.append(row_str)
                return f"[Excel: {filename}]\n" + "\n".join(lines[:200])

            # Word
            elif ext in ("docx", "doc"):
                doc = Document(io.BytesIO(raw))
                text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
                return f"[Word: {filename}]\n{text[:3000]}"

            # CSV / plain text
            elif ext in ("csv", "txt", "md"):
                return f"[{ext.upper()}: {filename}]\n{raw.decode('utf-8', errors='ignore')[:3000]}"

            # Image — use Groq vision to describe
            elif ext in ("png", "jpg", "jpeg", "gif", "webp"):
                b64 = base64.b64encode(raw).decode()
                mime = {"png": "image/png", "gif": "image/gif", "webp": "image/webp"}.get(ext, "image/jpeg")
                async with httpx.AsyncClient(timeout=30) as vc:
                    vr = await vc.post(
                        GROQ_URL,
                        headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
                        json={
                            "model": "meta-llama/llama-4-scout-17b-16e-instruct",
                            "max_tokens": 500,
                            "messages": [{
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": "Describe what this image shows in the context of a data/BI ticket. Include any numbers, charts, tables, or UI elements visible."},
                                    {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}}
                                ]
                            }]
                        }
                    )
                    vr_data = vr.json()
                    if "choices" not in vr_data:
                        return f"[Image: {filename}] (vision API error: {vr_data.get('error', {}).get('message', vr_data)})"
                    desc = vr_data["choices"][0]["message"]["content"]
                return f"[Image: {filename}]\n{desc}"

            else:
                return f"[Attachment: {filename}] (unsupported format)"
    except Exception as e:
        return f"[Attachment: {filename}] (could not read: {e})"


async def fetch_all_replies(comment_id, client, depth=1):
    """Recursively fetch sub-comments up to depth 3 to avoid infinite recursion."""
    if depth > 3:
        return []
    indent = "  " * depth
    lines = []
    try:
        resp = await client.get(
            f"{CLICKUP_BASE}/comment/{comment_id}/reply",
            headers={"Authorization": CLICKUP_API_KEY}
        )
        replies = resp.json().get("comments", [])
        for r in replies:
            user = (r.get("user") or {}).get("username", "unknown")
            text = extract_comment_text(r)
            reply_id = r.get("id", "")
            # Skip bot eval sub-comments — same filter as fetch_comments top-level.
            # Without this, bot gate reports posted as replies re-enter LLM context.
            if text and "🤖 **SubInspector" in text and (
                "Gate**" in text or "SCORE:" in text or "Auto-Completed" in text
            ):
                if reply_id:
                    sub_replies = await fetch_all_replies(reply_id, client, depth + 1)
                    lines.extend(sub_replies)
                continue
            if text:
                lines.append(f"{indent}↳ [{user}]: {text}")
            if reply_id:
                sub_replies = await fetch_all_replies(reply_id, client, depth + 1)
                lines.extend(sub_replies)
    except Exception:
        pass
    return lines


async def fetch_comments(task_id):
    """Fetch all comments. Returns (formatted_text_for_llm, raw_comment_list). Never raises."""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.get(
                f"{CLICKUP_BASE}/task/{task_id}/comment",
                headers={"Authorization": CLICKUP_API_KEY}
            )
            try:
                body = response.json()
            except Exception:
                print(f"[AGENT] fetch_comments: non-JSON response (HTTP {response.status_code})", flush=True)
                return "None", []
            comments = body.get("comments", []) if isinstance(body, dict) else []
            lines = []
            for c in comments:
                user = (c.get("user") or {}).get("username", "unknown")
                text = extract_comment_text(c)
                comment_id = c.get("id", "")
                # Skip SubInspector gate-check reports from the LLM context.
                # They contain ❌/✅ PASS/FAIL tables that confuse the model
                # when it re-evaluates the same ticket (causes 0/6 or 1/6 phantom scores).
                # Auto-generated closing notes (no "Gate**" / "SCORE:") are kept —
                # they represent real evidence of work completion.
                if text and "🤖 **SubInspector" in text and (
                    "Gate**" in text or "SCORE:" in text or "Auto-Completed" in text
                ):
                    print(f"[AGENT] Skipping bot eval comment from LLM context (user={user})", flush=True)
                    # Still append to lines for link extraction but mark it so we can strip later? No —
                    # just drop it entirely from LLM context; raw_comments keeps it for failure counting.
                    if comment_id:
                        try:
                            replies = await fetch_all_replies(comment_id, client, depth=1)
                            lines.extend(replies)  # keep any human replies inside the thread
                        except Exception:
                            pass
                    continue
                if text:
                    lines.append(f"[{user}]: {text}")
                if comment_id:
                    try:
                        replies = await fetch_all_replies(comment_id, client, depth=1)
                        lines.extend(replies)
                    except Exception:
                        pass
            full_text = "\n".join(lines) if lines else "None"
            print(f"[AGENT] fetch_comments: {len(comments)} top-level comments, {len(lines)} total lines, {len(full_text)} chars total", flush=True)
            return full_text, comments
    except Exception as e:
        print(f"[AGENT] fetch_comments failed entirely: {e}", flush=True)
        traceback.print_exc()
        return "None", []


async def evaluate_gate(gate, task, tier_override=None):
    task_id = task.get("id", "")
    # Process table-embeds BEFORE truncating:
    # • Large tables (>10 rows) → compact summary so they don't eat the char budget
    # • Small tables → human-readable text so BQ paths / KPI specs are visible to the LLM
    raw_description = task.get("description") or ""
    description = _process_table_embeds(raw_description)[:3500]
    assignees = ", ".join(a.get("username", "") for a in task.get("assignees", [])) or "None"
    list_name = (task.get("list") or {}).get("name", "")
    folder_name = (task.get("folder") or {}).get("name", "")

    # Read attachments in parallel — use return_exceptions so one bad attachment never kills the whole eval
    attachments = task.get("attachments", [])
    print(f"[AGENT] Attachments: {[a.get('title', a.get('file_name','?')) for a in attachments]}", flush=True)
    valid_attachments = [(a.get("url"), a.get("title", a.get("file_name", "attachment"))) for a in attachments[:5] if a.get("url")]
    if valid_attachments:
        try:
            attachment_results = await asyncio.gather(
                *[read_attachment(url, fn) for url, fn in valid_attachments],
                return_exceptions=True
            )
            attachment_info = "\n\n".join(
                r if isinstance(r, str) else f"[Attachment error: {r}]"
                for r in attachment_results
            )
        except Exception as e:
            print(f"[AGENT] Attachment gather failed: {e}", flush=True)
            attachment_info = "None"
    else:
        attachment_info = "None"

    # Fetch all subtasks in parallel — return_exceptions so one failing subtask doesn't abort
    subtask_stubs = task.get("subtasks", [])
    subtask_count = len(subtask_stubs)
    if subtask_stubs:
        try:
            async with httpx.AsyncClient(timeout=15) as st_client:
                async def _fetch_subtask(stub):
                    try:
                        r = await st_client.get(f"{CLICKUP_BASE}/task/{stub.get('id')}", headers={"Authorization": CLICKUP_API_KEY})
                        d = r.json()
                        return d if isinstance(d, dict) and d.get("id") else stub
                    except Exception:
                        return stub
                subtasks_raw = await asyncio.gather(
                    *[_fetch_subtask(s) for s in subtask_stubs],
                    return_exceptions=True
                )
                subtasks = [s if isinstance(s, dict) else stub for s, stub in zip(subtasks_raw, subtask_stubs)]
        except Exception as e:
            print(f"[AGENT] Subtask gather failed: {e}", flush=True)
            subtasks = subtask_stubs
        subtask_lines = []
        for s in subtasks:
            st_status = (s.get("status") or {}).get("status", "unknown")
            st_assignees = ", ".join(a.get("username", "") for a in (s.get("assignees") or [])) or "unassigned"
            line = f"- [{st_status}] {s.get('name', '')} (assignee: {st_assignees})"
            desc = (s.get("description") or "").strip()[:500]
            if desc:
                line += f"\n  Scope: {desc}"
            subtask_lines.append(line)
        subtask_info = "\n".join(subtask_lines)
    else:
        subtask_info = "None (no subtasks)"

    # Fetch custom fields — resolve dropdown option IDs to display names
    custom_fields = task.get("custom_fields", [])
    cf_lines = []
    for cf in custom_fields:
        name = cf.get("name", "")
        value = cf.get("value", "")
        field_type = cf.get("type", "")
        if not name or value in (None, "", [], {}):
            continue
        # Resolve dropdown/label option IDs to human-readable names
        if field_type in ("drop_down", "labels") and value:
            options = (cf.get("type_config") or {}).get("options", [])
            id_to_name = {str(o.get("id", "")): o.get("name", "") for o in options}
            if isinstance(value, list):
                resolved = [id_to_name.get(str(v), str(v)) for v in value]
                display = ", ".join(r for r in resolved if r)
            else:
                display = id_to_name.get(str(value), str(value))
        else:
            display = str(value)
        if display:
            cf_lines.append(f"- {name}: {display}")
    custom_fields_info = "\n".join(cf_lines) if cf_lines else "None"

    # Fetch comments and return both formatted text (for LLM) and raw list (for failure counter)
    comments_text, raw_comments = await fetch_comments(task_id)

    # Extract all notable links from the FULL comment text before capping,
    # so the LLM always knows about linked sheets/docs even if they're cut off.
    _url_re = re.compile(r'https?://\S+')
    _all_urls = _url_re.findall(comments_text)
    # Google Sheets — both docs.google.com/spreadsheets AND drive.google.com embeds
    _sheet_urls  = [u for u in _all_urls if
                    'docs.google.com/spreadsheet' in u
                    or 'sheets.google.com' in u
                    or ('drive.google.com' in u and ('spreadsheet' in u or 'gsheets' in u))]
    # Google Drive generic (file uploads, sheets without keyword in URL)
    _gdrive_urls = [u for u in _all_urls if 'drive.google.com' in u and u not in _sheet_urls]
    _gdoc_urls   = [u for u in _all_urls if 'docs.google.com/document' in u]
    _cu_doc_urls = [u for u in _all_urls if 'app.clickup.com' in u and '/docs/' in u]
    _gh_urls     = [u for u in _all_urls if 'github.com' in u and '/pull/' in u]
    _link_parts  = []
    if _sheet_urls:
        _link_parts.append("Google Sheets (validation/data evidence): " + ", ".join(_sheet_urls[:5]))
    if _gdrive_urls:
        _link_parts.append("Google Drive files (count as attached evidence): " + ", ".join(_gdrive_urls[:5]))
    if _gdoc_urls:
        _link_parts.append("Google Docs: " + ", ".join(_gdoc_urls[:3]))
    if _cu_doc_urls:
        _link_parts.append("ClickUp Docs: " + ", ".join(_cu_doc_urls[:3]))
    if _gh_urls:
        _link_parts.append("GitHub PRs: " + ", ".join(_gh_urls[:3]))
    links_section = ("Links found in comments (count as attached evidence):\n" + "\n".join(_link_parts)) if _link_parts else "None"
    print(f"[AGENT] Links extracted — sheets={len(_sheet_urls)} drive={len(_gdrive_urls)} docs={len(_gdoc_urls)} gh={len(_gh_urls)}", flush=True)

    # Cap sections — system prompt ~850 tokens, output ~1500 tokens,
    # links_section is tiny, leaving ~3500 tokens for the rest of the user message.
    #
    # IMPORTANT: use first-2000 + last-2000 so we always see BOTH the oldest
    # evidence (e.g. Feb validation sheet) AND the most recent closure notes.
    # A simple head-cap would cut off recent notes when there are many SI comments.
    _COMMENT_HEAD = 1200
    _COMMENT_TAIL = 1200
    if len(comments_text) > (_COMMENT_HEAD + _COMMENT_TAIL):
        comments_text_capped = (
            comments_text[:_COMMENT_HEAD]
            + "\n\n...[middle comments truncated — links/evidence above are extracted from ALL comments]...\n\n"
            + comments_text[-_COMMENT_TAIL:]
        )
    else:
        comments_text_capped = comments_text
    attachment_info_capped = attachment_info[:1500]
    subtask_info_capped    = subtask_info[:800]

    is_master = subtask_count > 0

    # Detect BI tickets in Python so the LLM doesn't have to guess
    task_name = task.get("name", "")
    bi_keywords = ["tableau", "power bi", "powerbi", "pbix", "dashboard", "workbook", "report"]
    is_bi = (
        task_name.lower().startswith("[bi]")
        or any(kw in task_name.lower() for kw in bi_keywords)
        or any(kw in (task.get("description") or "").lower()[:1000] for kw in bi_keywords)
    )
    bi_line = "Ticket Type: BI — use the BI-specific checklist, NOT the generic checklist.\n" if is_bi else "Ticket Type: DE (non-BI) — use the generic checklist.\n"
    print(f"[AGENT] BI detected: {is_bi}", flush=True)

    tier_line = f"Tier Override: {tier_override} (use this tier — do not infer)\n" if tier_override else ""
    user_message = (
        f"Gate: {gate}\n"
        f"{bi_line}"
        f"{tier_line}"
        f"Task: {task_name}\n"
        f"Status: {(task.get('status') or {}).get('status', '')}\n"
        f"Assignees: {assignees}\n"
        f"List: {list_name}\n"
        f"Folder: {folder_name}\n"
        f"Is Master Ticket: {'YES — has ' + str(subtask_count) + ' subtasks' if is_master else 'NO'}\n"
        f"NOTE: [table-embed:...] markers in the description are embedded data tables and count as evidence.\n"
        f"{links_section}\n\n"
        f"Description:\n{description}\n\n"
        f"Comments (closing notes, QA sign-offs, evidence, sub-comments):\n{comments_text_capped}\n\n"
        f"Attachments (actual content read):\n{attachment_info_capped}\n\n"
        f"Subtasks ({subtask_count} total — full details):\n{subtask_info_capped}\n\n"
        f"Custom Fields:\n{custom_fields_info}"
    )

    # Hard cap — CLOSURE system prompt ~1300 tokens + output 1500 tokens + user must stay under ~3200 tokens
    # 3200 tokens × 3.5 chars/token ≈ 11200 chars. Cap at 9000 for a safe margin.
    if len(user_message) > 9000:
        user_message = user_message[:9000] + "\n\n[TRUNCATED]"

    if not GROQ_API_KEY:
        raise ValueError("GROQ_API_KEY environment variable is not set")

    print(f"[AGENT] Sending to Groq — prompt size: {len(user_message)} chars", flush=True)

    class _RateLimitError(Exception):
        pass

    async def _call_groq(model: str) -> str:
        async with httpx.AsyncClient(timeout=60) as client:
            response = await client.post(
                GROQ_URL,
                headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
                json={
                    "model": model,
                    "temperature": 0,
                    "max_tokens": 1500,
                    "messages": [
                        {"role": "system", "content": get_system_prompt(gate)},
                        {"role": "user", "content": user_message}
                    ]
                }
            )
            if response.status_code == 429:
                raise _RateLimitError(f"rate limited on {model}")
            try:
                data = response.json()
            except Exception:
                raise ValueError(f"non-JSON response (HTTP {response.status_code}): {response.text[:400]}")
            if "choices" not in data:
                raise ValueError(f"Groq error on {model} (HTTP {response.status_code}): {data}")
            return data["choices"][0]["message"]["content"]

    # Strategy:
    # 1. Try llama-3.3-70b-versatile once — best quality, but strict rate limit (6k TPM free).
    # 2. On rate limit → immediately fall back to llama-3.1-8b-instant (20k TPM, no wait).
    # 3. Retry 8b-instant up to 3× with short waits if needed.
    last_error = None
    primary, fallback = "llama-3.3-70b-versatile", "llama-3.1-8b-instant"

    try:
        content = await _call_groq(primary)
        print(f"[AGENT] Groq OK — model={primary}", flush=True)
        return content, raw_comments, comments_text_capped
    except _RateLimitError as e:
        print(f"[AGENT] {primary} rate limited — switching to {fallback} immediately", flush=True)
        last_error = e
    except (httpx.TimeoutException, httpx.ConnectError) as e:
        print(f"[AGENT] {primary} network error: {e} — switching to {fallback}", flush=True)
        last_error = e
    except Exception as e:
        print(f"[AGENT] {primary} failed: {e} — switching to {fallback}", flush=True)
        last_error = e

    for attempt in range(3):
        try:
            content = await _call_groq(fallback)
            print(f"[AGENT] Groq OK — model={fallback} attempt={attempt+1}", flush=True)
            return content, raw_comments, comments_text_capped
        except _RateLimitError as e:
            wait_sec = 20 * (attempt + 1)
            print(f"[AGENT] {fallback} rate limited (attempt {attempt+1}/3) — waiting {wait_sec}s", flush=True)
            last_error = e
            await asyncio.sleep(wait_sec)
        except (httpx.TimeoutException, httpx.ConnectError) as e:
            print(f"[AGENT] {fallback} network error attempt {attempt+1}: {e}", flush=True)
            last_error = e
            await asyncio.sleep(5)
        except Exception as e:
            print(f"[AGENT] {fallback} failed attempt {attempt+1}: {e}", flush=True)
            last_error = e
            await asyncio.sleep(5)

    raise last_error or ValueError("Groq failed on all models")


async def post_comment(task_id, comment, reply_to_comment_id=None):
    """Post a comment on a task. If reply_to_comment_id is set, try to post as a reply inside
    that thread first; if the reply API call fails for any reason, fall back to a top-level
    comment so the message is never silently dropped.

    comment: str → sent as plain comment_text (for simple error/status messages)
             list → sent as rich-text comment block array (for formatted gate reports)
    """
    payload = {"comment": comment} if isinstance(comment, list) else {"comment_text": comment}
    async with httpx.AsyncClient(timeout=15) as client:
        if reply_to_comment_id:
            try:
                resp = await client.post(
                    f"{CLICKUP_BASE}/comment/{reply_to_comment_id}/reply",
                    headers={"Authorization": CLICKUP_API_KEY, "Content-Type": "application/json"},
                    json=payload
                )
                if resp.status_code < 300:
                    print(f"[AGENT] Reply posted to comment {reply_to_comment_id} (status {resp.status_code})", flush=True)
                    return
                print(f"[AGENT] Reply failed ({resp.status_code}) — falling back to top-level comment", flush=True)
            except Exception as e:
                print(f"[AGENT] Reply exception ({e}) — falling back to top-level comment", flush=True)

        # Post as a top-level comment (either no reply_to_comment_id, or reply failed above)
        resp = await client.post(
            f"{CLICKUP_BASE}/task/{task_id}/comment",
            headers={"Authorization": CLICKUP_API_KEY, "Content-Type": "application/json"},
            json=payload
        )
        if resp.status_code < 300:
            print(f"[AGENT] Top-level comment posted on task {task_id} (status {resp.status_code})", flush=True)
        else:
            print(f"[AGENT] ⚠️ Top-level comment also failed ({resp.status_code}): {resp.text[:200]}", flush=True)


async def revert_status(task_id, status) -> bool:
    """Returns True if the status was successfully changed, False otherwise."""
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.put(
            f"{CLICKUP_BASE}/task/{task_id}",
            headers={"Authorization": CLICKUP_API_KEY, "Content-Type": "application/json"},
            json={"status": status}
        )
        if resp.status_code < 300:
            print(f"[AGENT] Status reverted to '{status}' on task {task_id} (HTTP {resp.status_code})", flush=True)
            return True
        else:
            print(f"[AGENT] ⚠️ Revert failed (HTTP {resp.status_code}): {resp.text[:300]}", flush=True)
            return False


def format_comment(gate, content, score, passed, prior_failures=0, reverted_to=None, advisory=False):
    """Parse LLM output and return a ClickUp rich-text comment block array.

    advisory=True: compact report for out-of-scope tasks — no status revert,
    no escalation, no next-steps. Just the gate result + per-check gaps.
    """

    # ── extract pieces from LLM response ──────────────────────────────────
    tier_match    = re.search(r"TIER:\s*(.+)", content)
    summary_match = re.search(r"SUMMARY:\s*(.+)", content)
    checks_match  = re.search(r"CHECKS:\n(.*?)(?=\nSUMMARY:|\nMASTER TICKET:|$)", content, re.DOTALL)
    master_match  = re.search(r"(MASTER TICKET:.*)", content, re.DOTALL)

    tier_line    = tier_match.group(1).strip()    if tier_match    else "—"
    summary      = summary_match.group(1).strip() if summary_match else ""
    checks_table = checks_match.group(1).strip()  if checks_match  else ""
    master_block = master_match.group(1).strip()  if master_match  else ""

    result_emoji = "✅" if passed else "❌"
    result_word  = "PASS" if passed else "FAIL"

    # ── header ─────────────────────────────────────────────────────────────
    blocks = [
        {"text": f"🤖 SubInspector — {gate} Gate\n", "attributes": {"bold": True}},
    ]
    if advisory:
        blocks.append({"text": "🔍 Advisory mode — outside Instant Hydration folder. No status changes made.\n"})
    blocks += [
        {"text": "\n"},
        {"text": "🏷 Tier: ", "attributes": {"bold": True}},
        {"text": f"{tier_line}\n"},
        {"text": "📊 Score: ", "attributes": {"bold": True}},
        {"text": f"{score}/6  |  {result_emoji} {result_word}\n"},
    ]
    if reverted_to:
        blocks += [
            {"text": "🔁 Status reverted to: ", "attributes": {"bold": True}},
            {"text": f"{reverted_to}\n"},
        ]
    blocks.append({"text": "\n"})

    # ── checks table ────────────────────────────────────────────────────────
    if checks_table:
        blocks += [
            {"text": "Gate Checks\n", "attributes": {"bold": True}},
            {"text": "\n"},
            {"text": checks_table + "\n"},
            {"text": "\n"},
        ]

    # ── summary ─────────────────────────────────────────────────────────────
    if summary:
        blocks += [
            {"text": "📝 Summary: ", "attributes": {"bold": True}},
            {"text": f"{summary}\n"},
        ]

    # ── next steps / escalation ─────────────────────────────────────────────
    # Skipped in advisory mode — no enforcement, just per-check gap feedback.
    if not passed and not advisory:
        blocks.append({"text": "\n"})
        if prior_failures == 0:
            blocks += [
                {"text": "💡 Next Steps\n", "attributes": {"bold": True}},
                {"text": "  -  Fix every ❌ check listed above\n"},
                {"text": "  -  Once updated, trigger a SubInspector re-check to re-evaluate\n"},
            ]
        elif prior_failures == 1:
            blocks += [
                {"text": "⚠️ 2nd Failure — BA Lead Consult Required\n", "attributes": {"bold": True}},
                {"text": "  -  This ticket has failed SubInspector gate checks twice\n"},
                {"text": "  -  Please discuss with @Komal Saraogi before making further changes\n"},
                {"text": "  -  Fix all ❌ checks above, then trigger a re-check to retry\n"},
            ]
        else:
            blocks += [
                {"text": f"🚨 Repeated Failure ({prior_failures + 1} total) — Enforcement Suspended\n", "attributes": {"bold": True}},
                {"text": "  -  @Komal Saraogi — manual review required before this ticket can proceed\n"},
                {"text": "  -  Automatic gate enforcement is paused; the team must resolve this manually\n"},
            ]

    # ── master ticket scope analysis ────────────────────────────────────────
    if master_block:
        blocks += [
            {"text": "\n"},
            {"text": "📋 Master Ticket Scope Analysis\n", "attributes": {"bold": True}},
            {"text": "\n"},
            {"text": master_block + "\n"},
        ]

    return blocks


async def count_subinspector_failures(task_id, gate=None, raw_comments=None):
    """Count prior SubInspector FAIL comments for a specific gate.

    gate: "INTAKE" | "PRE-EXECUTION" | "CLOSURE" (or None to count across all gates — not recommended).
    Counting per-gate avoids escalating to BA lead on a CLOSURE check just because
    the ticket previously failed PRE-EXECUTION twice.
    """
    try:
        if raw_comments is None:
            async with httpx.AsyncClient(timeout=15) as client:
                resp = await client.get(
                    f"{CLICKUP_BASE}/task/{task_id}/comment",
                    headers={"Authorization": CLICKUP_API_KEY}
                )
                raw_comments = resp.json().get("comments", [])
        count = 0
        for c in raw_comments:
            text = extract_comment_text(c)
            # Match only the bot's own structured failure comments by their signature header.
            # If gate is specified, only count failures from the same gate so that
            # repeated failures at one gate don't inflate the failure counter at another.
            gate_marker = f"SubInspector — {gate} Gate" if gate else "SubInspector"
            if gate_marker in text and "❌" in text:
                count += 1
        return count
    except Exception:
        return 0


# Checks whose gap can be filled by SI writing content.
# "evidence" and "open subtask" require real human action — not auto-fixable.
_AUTO_FIXABLE_CHECK_KEYWORDS = [
    "acceptance criteria",
    "qa sign-off",
    "sign-off",
    "stakeholder notified",
    "stakeholder",
    "documentation updated",
    "documentation",
]


def _can_auto_complete(score: int, content: str) -> tuple[bool, list[str]]:
    """
    Returns (can_auto_complete, failing_check_names).
    Auto-complete triggers when score >= 4 AND every failing check is a soft
    formality SI can write (closing note / acceptance confirmation / stakeholder
    mention / docs N/A). Score floor of 4 prevents auto-closing tickets that
    have real substantive gaps (missing evidence, open subtasks, etc.).
    """
    if score == 6:
        return False, []  # already passing
    if score < 4:
        return False, []  # too many gaps — require human to fix
    failing = re.findall(r'\|\s*\d+\s*\|\s*([^|]+?)\s*\|\s*❌\s*FAIL', content)
    if not failing:
        return False, []
    can_fix = all(
        any(kw in check.lower() for kw in _AUTO_FIXABLE_CHECK_KEYWORDS)
        for check in failing
    )
    return can_fix, failing


def _resolve_stakeholder(task: dict) -> str:
    """
    Returns the ticket creator's username as the stakeholder to notify.
    If creator info is missing, returns an empty string (no @mention added).
    """
    creator = task.get("creator") or {}
    return creator.get("username", "")


async def generate_auto_closing_note(task: dict, comments_text: str) -> str:
    """
    Use Groq to draft a closing note from ticket context.
    Uses the small/fast 8b model — output is short (~300 tokens).
    """
    import datetime
    today       = datetime.date.today().strftime("%b %d, %Y")
    task_name   = task.get("name", "")
    assignees   = ", ".join(a.get("username", "") for a in task.get("assignees", [])) or "team"
    description = _process_table_embeds(task.get("description") or "")[:1500]
    stakeholder = _resolve_stakeholder(task)

    prompt = f"""You are writing a professional closing note for a ClickUp ticket on behalf of the team lead.

Ticket: {task_name}
Assignees: {assignees}
Stakeholder to notify: {stakeholder}
Today: {today}
Description (summary):
{description}

Recent comments / work evidence:
{comments_text[:1800]}

Write a concise closing note with EXACTLY this structure:
✅ Closure Notes — {today}
---
**What was delivered:**
- [bullet based on description/comments — factual, no invention]
- [bullet]

**Evidence:**
- [reference any validation, PR, query, or sheet mentioned in comments]
- All acceptance criteria addressed ✓
- No open subtasks or blockers ✓

Moving ticket to **Done**. 🎉

Important rules:
- {"@mention " + stakeholder + " at the end to notify them the ticket is done" if stakeholder else "Do NOT @mention anyone — creator unknown"}
- Only include facts present in the description or comments — do NOT invent
- Keep bullets short (one line each)
- No extra commentary outside the format above"""

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                GROQ_URL,
                headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
                json={
                    "model": "llama-3.1-8b-instant",
                    "temperature": 0,
                    "max_tokens": 400,
                    "messages": [{"role": "user", "content": prompt}]
                }
            )
            data = resp.json()
            if "choices" in data:
                return data["choices"][0]["message"]["content"].strip()
    except Exception as e:
        print(f"[AGENT] generate_auto_closing_note failed: {e}", flush=True)
    return ""


async def fetch_comment_text_from_api(comment_id: str) -> str:
    """Fetch a single comment's text directly from the ClickUp API as a fallback."""
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(
                f"{CLICKUP_BASE}/comment/{comment_id}",
                headers={"Authorization": CLICKUP_API_KEY}
            )
            if resp.status_code == 200:
                body = resp.json()
                text = extract_comment_text(body) or extract_comment_text(body.get("comment") or {})
                if text:
                    return text
    except Exception as e:
        print(f"[AGENT] fetch_comment_text_from_api({comment_id}) failed: {e}", flush=True)
    return ""


async def process_webhook(payload):
    event = payload.get("event")
    task_id = payload.get("task_id")
    history_items = payload.get("history_items", [])

    if not task_id or not event:
        return

    print(f"[AGENT] Webhook: event={event} task_id={task_id} history_items_count={len(history_items)}", flush=True)
    if history_items:
        # Log structure so we can debug comment text extraction misses
        import json
        sample = history_items[0]
        print(f"[AGENT] history_items[0] keys={list(sample.keys())}", flush=True)
        comment_preview = sample.get("comment") or (sample.get("data") or {}).get("comment") or {}
        print(f"[AGENT] comment_obj keys={list(comment_preview.keys()) if isinstance(comment_preview, dict) else type(comment_preview)}", flush=True)

    # Bot-account loop prevention — two cases:
    # 1. taskStatusUpdated from bot → always skip (prevents revert → re-evaluate loop)
    # 2. taskCommentPosted from bot → skip UNLESS it's a short bare trigger phrase
    #    (≤ 100 chars). The bot's evaluation reports are hundreds of chars; a human
    #    typing /si check is < 20 chars. This makes loops structurally impossible:
    #    even if the LLM somehow outputs the trigger phrase, the comment length will
    #    be >> 100 chars and will be skipped.
    if history_items:
        actor_id = str((history_items[0].get("user") or {}).get("id", ""))
        if actor_id == BOT_USER_ID:
            if event == "taskStatusUpdated":
                print(f"[AGENT] Skipping — taskStatusUpdated from bot account", flush=True)
                return
            if event == "taskCommentPosted":
                item = history_items[0]
                comment_obj = (
                    item.get("comment")
                    or (item.get("data") or {}).get("comment")
                    or {}
                )
                raw_text = (extract_comment_text(comment_obj) or extract_comment_text(item) or "").strip()

                # If text extraction from webhook payload failed, try fetching via API
                if not raw_text:
                    comment_id = (comment_obj.get("id") if isinstance(comment_obj, dict) else None) or item.get("id") or ""
                    if comment_id:
                        print(f"[AGENT] Text extraction from payload failed — fetching comment {comment_id} from API", flush=True)
                        raw_text = await fetch_comment_text_from_api(comment_id)

                print(f"[AGENT] Bot-account comment text: {repr(raw_text[:120])}", flush=True)

                # Only skip if we have text AND it's NOT a short trigger.
                # If text is empty (extraction failed entirely), let it through —
                # determine_gate will check for trigger and gate=None → skip safely.
                if raw_text and not (_is_trigger(raw_text) and len(raw_text) <= 100):
                    print(f"[AGENT] Skipping — bot account comment (len={len(raw_text)})", flush=True)
                    return
                elif not raw_text:
                    print(f"[AGENT] Bot account comment — could not extract text, letting through to determine_gate", flush=True)

    try:
        task = await fetch_task(task_id)
    except Exception as e:
        print(f"[AGENT] fetch_task failed for {task_id}: {e}", flush=True)
        return

    folder_id = str((task.get("folder") or {}).get("id", ""))
    list_id   = str((task.get("list")   or {}).get("id", ""))
    space_id  = str((task.get("space")  or {}).get("id", ""))

    # Three-level scope check (OR logic):
    #   folder.id → catches regular sprint/project tasks
    #   list.id   → fallback for tasks whose folder.id = "none" (list directly in space)
    #   space.id  → broadest: catches master/epic tickets in any folder within the space
    in_scope = (
        folder_id in ENFORCEMENT_FOLDERS
        or list_id in ENFORCEMENT_FOLDERS
        or (bool(ENFORCEMENT_SPACES) and space_id in ENFORCEMENT_SPACES)
    )
    in_advisory = (
        folder_id in ADVISORY_FOLDERS
        or list_id in ADVISORY_FOLDERS
        or (bool(ADVISORY_SPACES) and space_id in ADVISORY_SPACES)
    )
    print(f"[AGENT] Folder: {folder_id} | List: {list_id} | Space: {space_id} | Enforcement: {in_scope} | Advisory: {in_advisory}", flush=True)

    if not in_scope and not in_advisory:
        # None of folder / list / space matched — skip entirely.
        print(f"[AGENT] Skipping — folder/list/space not in any enforcement or advisory set", flush=True)
        return

    if not in_scope:
        # Advisory folder: comment only, no status changes, no escalation.
        # taskStatusUpdated → gate check for new status (no revert)
        # taskCreated       → INTAKE check
        # taskCommentPosted → gate check on /si check trigger only (checked after determine_gate)
        if event not in ("taskCommentPosted", "taskCreated", "taskStatusUpdated"):
            print(f"[AGENT] Skipping — advisory folder, unhandled event={event}", flush=True)
            return
        print(f"[AGENT] Advisory mode (event={event})", flush=True)

    advisory_mode = not in_scope

    status = (task.get("status") or {}).get("status", "")
    previous_status = ""
    if history_items:
        before = history_items[0].get("before") or {}
        previous_status = (before.get("status", "") if isinstance(before, dict) else "") or ""

    gate, is_dry_run, trigger_comment_id, tier_override = determine_gate(event, status, history_items)

    # For advisory taskCommentPosted, gate=None means no /si check trigger — skip.
    # For advisory taskCreated, determine_gate always returns INTAKE so gate won't be None.
    if advisory_mode and not gate:
        print(f"[AGENT] Out-of-scope task — no /si check trigger found, skipping", flush=True)
        return

    # Advisory mode: always dry-run, never touch status or fields.
    if advisory_mode:
        is_dry_run = True
        print(f"[AGENT] Advisory mode — enforcement disabled, reporting only", flush=True)

    # For CLOSURE gate with no previous_status (e.g. manual /si check), use the
    # revert map so enforcement still works even without a webhook history entry.
    if gate == "CLOSURE" and not previous_status and not advisory_mode:
        previous_status = CLOSURE_REVERT_MAP.get(status.lower(), "")
        if previous_status:
            print(f"[AGENT] No previous_status in payload — CLOSURE_REVERT_MAP: '{status}' → '{previous_status}'", flush=True)
    if gate == "PRE-EXECUTION" and not previous_status and not advisory_mode:
        previous_status = PRE_EXEC_REVERT_MAP.get(status.lower(), "open")
        print(f"[AGENT] No previous_status in payload — PRE_EXEC_REVERT_MAP: '{status}' → '{previous_status}'", flush=True)
    print(f"[AGENT] Gate: {gate} | Status: {status} | Advisory: {advisory_mode} | Reply to: {trigger_comment_id} | Tier override: {tier_override}", flush=True)

    if not gate:
        print(f"[AGENT] No gate matched — skipping", flush=True)
        return

    try:
        content, raw_comments, comments_text_for_note = await evaluate_gate(gate, task, tier_override=tier_override)
    except Exception as e:
        err_type = type(e).__name__
        err_msg = str(e)[:300]
        print(f"[AGENT] evaluate_gate failed — {err_type}: {err_msg}", flush=True)
        traceback.print_exc()
        await post_comment(
            task_id,
            f"⚠️ SubInspector — {gate} gate check failed.\n`{err_type}: {err_msg}`",
            reply_to_comment_id=trigger_comment_id
        )
        return

    # Strip any trigger phrase the LLM may have hallucinated (loop prevention layer 3).
    for _tp in _TRIGGER_PATTERNS:
        content = _tp.sub("[re-check command]", content)

    # BQ path override: if check #3 ❌ FAIL but a BQ path exists in the description,
    # the LLM is wrong (new-build/ingestion ticket with target paths). Fix it in Python.
    _desc_for_bq = _process_table_embeds(task.get("description", "") or "")
    content = _fix_bq_check_false_fail(content, _desc_for_bq)

    # Count actual ✅ PASS entries — more reliable than trusting the LLM's stated SCORE.
    checks_match = re.search(r"CHECKS:\n(.*?)(?=\nSUMMARY:|\nMASTER TICKET:|$)", content, re.DOTALL)
    if checks_match:
        score = str(checks_match.group(1).count("✅ PASS"))
    else:
        score_match = re.search(r"SCORE:\s*(\d+)/6", content, re.IGNORECASE)
        score = score_match.group(1) if score_match else "0"

    passed = int(score.strip()) == 6

    # ── AUTO-COMPLETE ──────────────────────────────────────────────────────────
    # When CLOSURE gate scores 5/6 and the one failing check is a soft formality
    # (closing note, stakeholder mention, docs N/A), SI writes the missing
    # content, posts it, and moves the ticket to complete automatically.
    # Disabled in advisory mode — never modify tasks outside ENFORCEMENT_FOLDERS.
    if not advisory_mode and gate == "CLOSURE" and not passed and status.lower() not in ("complete", "done"):
        can_fix, failing_checks = _can_auto_complete(int(score.strip()), content)
        if can_fix:
            print(f"[AGENT] Auto-complete triggered — soft gaps: {failing_checks}", flush=True)
            closing_note = await generate_auto_closing_note(task, comments_text_for_note)
            if closing_note:
                for _tp in _TRIGGER_PATTERNS:
                    closing_note = _tp.sub("[re-check command]", closing_note)
                await post_comment(
                    task_id,
                    f"🤖 **SubInspector — Auto-Generated Closing Note**\n\n{closing_note}\n\n"
                    f"_Auto-generated by SubInspector based on ticket context and comments._",
                    reply_to_comment_id=trigger_comment_id
                )
                moved = await revert_status(task_id, "complete")
                status_line = "✅ Ticket moved to **complete**." if moved else "⚠️ Could not update status — please move manually."
                await post_comment(
                    task_id,
                    f"🤖 **SubInspector — Auto-Completed** | Score {score}/6\n\n"
                    f"The only gap (`{'`, `'.join(failing_checks)}`) was a formality SI could fill.\n"
                    f"Closing note posted above. {status_line}",
                    reply_to_comment_id=trigger_comment_id
                )
                print(f"[AGENT] Auto-complete done for {task_id}", flush=True)
                return
            print(f"[AGENT] Auto-complete: note generation failed — falling back to normal FAIL flow", flush=True)
    # ── END AUTO-COMPLETE ──────────────────────────────────────────────────────

    # Failure escalation and status revert are enforcement-only actions.
    # In advisory mode (out-of-scope tasks) we skip both.
    prior_failures = 0
    if not passed and not advisory_mode:
        prior_failures = await count_subinspector_failures(task_id, gate=gate, raw_comments=raw_comments)
        if prior_failures >= 2:
            print(f"[AGENT] Anti-loop triggered after {prior_failures + 1} failures — escalating to BA lead", flush=True)

    can_revert = not passed and not is_dry_run and bool(previous_status) and not advisory_mode
    reverted_to = None
    if can_revert:
        print(f"[AGENT] Reverting status to: {previous_status}", flush=True)
        success = await revert_status(task_id, previous_status)
        reverted_to = previous_status if success else None
        if not success:
            print(f"[AGENT] ⚠️ Revert failed — comment will NOT claim status was changed", flush=True)

    comment = format_comment(gate, content, score, passed, prior_failures, reverted_to=reverted_to, advisory=advisory_mode)

    await post_comment(task_id, comment, reply_to_comment_id=trigger_comment_id)


# ── Backfill / missed-ticket scan ─────────────────────────────────────────────

async def has_gate_comment(task_id: str, gate: str, raw_comments=None) -> bool:
    """Return True if SubInspector has already posted a gate comment for this gate."""
    try:
        if raw_comments is None:
            async with httpx.AsyncClient(timeout=15) as client:
                resp = await client.get(
                    f"{CLICKUP_BASE}/task/{task_id}/comment",
                    headers={"Authorization": CLICKUP_API_KEY}
                )
                raw_comments = resp.json().get("comments", [])
        gate_marker = f"SubInspector — {gate} Gate"
        for c in raw_comments:
            text = extract_comment_text(c)
            if gate_marker in text:
                return True
        return False
    except Exception:
        return False


async def fetch_folder_tasks(folder_id: str) -> list:
    """Fetch all tasks across every list in a folder, paginating automatically."""
    all_tasks = []
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            resp = await client.get(
                f"{CLICKUP_BASE}/folder/{folder_id}/list",
                headers={"Authorization": CLICKUP_API_KEY}
            )
            lists = resp.json().get("lists", [])
            print(f"[SCAN] Found {len(lists)} lists in folder {folder_id}", flush=True)

            for lst in lists:
                list_id   = lst.get("id")
                list_name = lst.get("name", "?")
                page      = 0
                list_count = 0
                while True:
                    tr = await client.get(
                        f"{CLICKUP_BASE}/list/{list_id}/task",
                        headers={"Authorization": CLICKUP_API_KEY},
                        params={
                            "include_closed": "true",
                            "subtasks": "true",
                            "page": page,
                        }
                    )
                    page_tasks = tr.json().get("tasks", [])
                    all_tasks.extend(page_tasks)
                    list_count += len(page_tasks)
                    if len(page_tasks) < 100:   # ClickUp max 100/page
                        break
                    page += 1
                print(f"[SCAN]   List '{list_name}' ({list_id}): {list_count} tasks", flush=True)
    except Exception as e:
        print(f"[SCAN] fetch_folder_tasks failed: {e}", flush=True)
    return all_tasks


async def scan_and_backfill(folder_id: str = None, dry_run: bool = False) -> dict:
    """
    Scan every task in the IH folder. For each task:
      1. Determine which gate should have fired for the current status.
      2. Check whether SubInspector already posted that gate's comment.
      3. If not — evaluate and post (no status revert for backfills).

    dry_run=True  → identify missed tickets only, don't post anything.
    Returns a summary dict.
    """
    target_folder = folder_id or ENFORCEMENT_FOLDERS[0]
    print(f"[SCAN] Starting backfill — folder={target_folder} dry_run={dry_run}", flush=True)

    tasks = await fetch_folder_tasks(target_folder)
    print(f"[SCAN] Total tasks to evaluate: {len(tasks)}", flush=True)

    results = {
        "scanned": 0,
        "already_covered": 0,
        "missed": 0,
        "posted": 0,
        "errors": 0,
    }
    missed_list = []

    for task in tasks:
        task_id   = task.get("id", "")
        task_name = task.get("name", "")
        status    = (task.get("status") or {}).get("status", "").lower()
        results["scanned"] += 1

        # Determine the expected gate for the current status
        if any(s in status for s in CLOSURE_STATUSES):
            expected_gate = "CLOSURE"
        elif any(s in status for s in PRE_EXEC_STATUSES):
            expected_gate = "PRE-EXECUTION"
        else:
            expected_gate = "INTAKE"

        already = await has_gate_comment(task_id, expected_gate)
        if already:
            results["already_covered"] += 1
            continue

        results["missed"] += 1
        missed_list.append({"id": task_id, "name": task_name, "status": status, "gate": expected_gate})
        print(f"[SCAN] MISSED — {task_id} | '{task_name[:60]}' | status={status} | gate={expected_gate}", flush=True)

        if dry_run:
            continue

        try:
            full_task = await fetch_task(task_id)
            content, raw_comments, _ = await evaluate_gate(expected_gate, full_task)

            # Strip any trigger phrase the LLM may have hallucinated
            for _tp in _TRIGGER_PATTERNS:
                content = _tp.sub("[re-check command]", content)

            # BQ path override (backfill path)
            _desc_bq = _process_table_embeds(full_task.get("description", "") or "")
            content = _fix_bq_check_false_fail(content, _desc_bq)

            checks_match = re.search(r"CHECKS:\n(.*?)(?=\nSUMMARY:|\nMASTER TICKET:|$)", content, re.DOTALL)
            if checks_match:
                score = str(checks_match.group(1).count("✅ PASS"))
            else:
                score_match = re.search(r"SCORE:\s*(\d+)/6", content, re.IGNORECASE)
                score = score_match.group(1) if score_match else "0"

            passed        = int(score.strip()) == 6
            prior_failures = await count_subinspector_failures(task_id, gate=expected_gate, raw_comments=raw_comments)

            # Backfill comments never revert status — ticket may have moved on since
            comment = format_comment(expected_gate, content, score, passed, prior_failures, reverted_to=None)
            await post_comment(task_id, comment)
            results["posted"] += 1
            print(f"[SCAN] Posted {expected_gate} gate on {task_id} — score={score}/6 passed={passed}", flush=True)

            await asyncio.sleep(1)   # gentle rate-limit buffer between posts

        except Exception as e:
            results["errors"] += 1
            print(f"[SCAN] Error on {task_id}: {e}", flush=True)

    results["missed_tickets"] = missed_list
    print(f"[SCAN] Done — {results}", flush=True)
    return results
